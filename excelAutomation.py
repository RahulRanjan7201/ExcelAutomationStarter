import openpyxl

workbook = openpyxl.load_workbook("Employees.xlsx")
print(workbook.properties)
print(workbook.sheetnames)
print(workbook.active)
sheet = workbook['EmployeeData']
print(sheet)
for i in sheet.values :
    print(i)

print(sheet['B7'].value)
cell = sheet['B9']
print(cell.row)
print(cell.column)
cell = sheet['B2']
cell.value = 'David'
#create workbook 
workbook.create_sheet("TestSheet")
workbook.save("Employees.xlsx")
sheet = workbook['TestSheet']
workbook.remove(sheet)
workbook.save("Employees.xlsx")
print(sheet.title)
